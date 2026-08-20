"""
generador_excel.py — Matriz de Competencias plan 2025
Construye el .xlsx completo con openpyxl.Workbook() desde cero.
No abre ni lee data/plantilla_matriz.xlsx en ningún momento.
"""
import sqlite3
import argparse
from collections import OrderedDict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

DB_PATH   = Path("data/sistema.db")
OUT_DIR   = Path("data/output")
LOGO_PATH = Path("data/uv_logo_nuevo.png")

# Colores institucionales (mismos que generador_word.py)
HEX = {
    "licenciatura": "1F4E79",
    "titulo":       "375623",
    "sello_uv":     "843C0C",
}
HEX_CLARO = {
    "licenciatura": "DEEAF1",
    "titulo":       "E2EFDA",
    "sello_uv":     "FCE4D6",
}
HEX_GRIS   = "D6DCE4"
HEX_BLANCO = "FFFFFF"

# Tonos por competencia individual (más oscuro → más claro dentro de cada tipo)
HEX_COMP = {
    # Licenciatura — azules
    "CL1": "BDD7EE",
    "CL2": "DEEAF1",
    # Título — verdes
    "CE1": "C6E0B4",
    "CE2": "E2EFDA",
    # Sello UV — naranjas
    "CG1": "F4B183",
    "CG2": "F8CBAD",
    "CG3": "FAD9C1",
    "CG4": "FCE4D6",
}

LABEL = {
    "licenciatura": "Competencias de Licenciatura",
    "titulo":       "Competencias Específicas del Título",
    "sello_uv":     "Competencias Genéricas Sello UV",
}

# ── Helpers de estilo ────────────────────────────────────────────────

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, size=10, color="000000", italic=False):
    return Font(name="Calibri", size=size, bold=bold,
                color=color, italic=italic)

def _align(h="center", v="center", wrap=False, rotation=0):
    return Alignment(horizontal=h, vertical=v,
                     wrap_text=wrap, text_rotation=rotation)

def _border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _write(ws, row, col, value, font=None, fill=None,
           align=None, border=None):
    c = ws.cell(row=row, column=col)
    c.value = value
    if font:   c.font   = font
    if fill:   c.fill   = fill
    if align:  c.alignment = align
    if border: c.border = border
    return c

# ── BD ───────────────────────────────────────────────────────────────

def _conectar():
    if not DB_PATH.exists():
        raise FileNotFoundError(f"BD no encontrada: {DB_PATH}")
    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = None
    return conn

def _obtener_datos():
    conn = _conectar()

    ras = conn.execute("""
        SELECT ra.id, ra.codigo_completo, nd.codigo_nivel, ra.codigo_ra,
               c.id, c.codigo, c.tipo
        FROM resultados_aprendizaje ra
        JOIN niveles_dominio nd ON nd.id = ra.nivel_dominio_id
        JOIN competencias c ON c.id = nd.competencia_id
        WHERE c.tipo != 'desconocido'
        ORDER BY
            CASE c.tipo
                WHEN 'licenciatura' THEN 0
                WHEN 'titulo'       THEN 1
                WHEN 'sello_uv'     THEN 2
            END,
            c.codigo,
            nd.codigo_nivel,
            ra.codigo_ra
    """).fetchall()

    asigs = conn.execute("""
        SELECT id, codigo, nombre, semestre, creditos
        FROM asignaturas
        ORDER BY semestre, codigo
    """).fetchall()

    tribs = set(conn.execute(
        "SELECT asignatura_id, ra_id FROM tributaciones"
    ).fetchall())

    conn.close()
    return ras, asigs, tribs

def _nivel_academico(semestre):
    if semestre is None: return "N1"
    s = int(semestre)
    if s <= 4:  return "N1"
    if s <= 8:  return "N2"
    return "N3"

# ── Generador ────────────────────────────────────────────────────────

def _subtones_ciclo(base_hex, n):
    """Genera n sub-tonos progresivos del color base (de más claro a más oscuro)."""
    r = int(base_hex[:2], 16); g = int(base_hex[2:4], 16); b = int(base_hex[4:], 16)
    tones = []
    for i in range(n):
        t = 0.12 + (i / max(n - 1, 1)) * 0.35 if n > 1 else 0.22
        nr = int(255 + (r - 255) * t); ng = int(255 + (g - 255) * t); nb = int(255 + (b - 255) * t)
        tones.append(f"{nr:02X}{ng:02X}{nb:02X}")
    return tones


def generar_matriz(salida=None, por_ciclo=False):
    """Genera la matriz de competencias.

    por_ciclo=True: cada nivel de dominio dentro de una competencia recibe
    un sub-tono distinto del color base de esa competencia.
    por_ciclo=False (default): usa HEX_COMP (comportamiento original).
    """
    ras, asigs, tribs = _obtener_datos()

    if not ras:
        print("⚠ Sin RAs válidos en la BD.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Matriz de Competencias"

    # ra_list: (ra_id, codigo_completo, nivel_dominio, codigo_ra, comp_codigo, comp_tipo)
    ra_list = [(r[0], r[1], r[2], r[3], r[5], r[6]) for r in ras]

    # Construir mapa (comp_cod, nivel) → color de fondo para modo por_ciclo
    ciclo_color = {}
    if por_ciclo:
        from collections import defaultdict
        comp_niveles = defaultdict(list)
        for _, _, nivel, _, comp_cod, comp_tipo in ra_list:
            if nivel not in comp_niveles[(comp_cod, comp_tipo)]:
                comp_niveles[(comp_cod, comp_tipo)].append(nivel)
        for (comp_cod, comp_tipo), niveles in comp_niveles.items():
            base = HEX[comp_tipo]
            tones = _subtones_ciclo(base, len(niveles))
            for i, nivel in enumerate(sorted(niveles)):
                ciclo_color[(comp_cod, nivel)] = tones[i]

    def _bg_ra(comp_cod, comp_tipo, nivel):
        if por_ciclo:
            return ciclo_color.get((comp_cod, nivel), HEX_CLARO[comp_tipo])
        return HEX_COMP.get(comp_cod, HEX_CLARO[comp_tipo])

    # Agrupar RAs por competencia (orden preservado)
    comp_ras = OrderedDict()
    for ra_id, ccomp, nivel, cod_ra, comp_cod, comp_tipo in ra_list:
        comp_ras.setdefault((comp_cod, comp_tipo), []).append(ra_id)

    # Agrupar competencias por tipo
    tipo_comps = OrderedDict()
    for (comp_cod, comp_tipo), ra_ids in comp_ras.items():
        tipo_comps.setdefault(comp_tipo, []).append((comp_cod, ra_ids))

    # ── Columnas fijas ────────────────────────────────────────────────
    COL_NIVEL  = 1
    COL_SEM    = 2
    COL_COD    = 3
    COL_NOM    = 4
    COL_CRED   = 5
    COL_TOT    = 6
    COL_LIC    = 7
    COL_TIT    = 8
    COL_SELLO  = 9
    COL_RA_INI = 10

    # Mapear ra_id → columna
    ra_to_col = {}
    col = COL_RA_INI
    for ra_id, *_ in ra_list:
        ra_to_col[ra_id] = col
        col += 1

    total_cols = COL_RA_INI + len(ra_list) - 1

    # Rangos de columna por tipo (para fórmulas COUNTIF)
    tipo_col_range = {}
    for comp_tipo, comp_list in tipo_comps.items():
        cols = []
        for _, ra_ids in comp_list:
            cols += [ra_to_col[r] for r in ra_ids]
        if cols:
            tipo_col_range[comp_tipo] = (min(cols), max(cols))

    # ── Anchos de columna ─────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(COL_NIVEL)].width = 4
    ws.column_dimensions[get_column_letter(COL_SEM)].width   = 4
    ws.column_dimensions[get_column_letter(COL_COD)].width   = 12
    ws.column_dimensions[get_column_letter(COL_NOM)].width   = 36
    ws.column_dimensions[get_column_letter(COL_CRED)].width  = 5
    for c in [COL_TOT, COL_LIC, COL_TIT, COL_SELLO]:
        ws.column_dimensions[get_column_letter(c)].width = 5
    for c in range(COL_RA_INI, total_cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = 4

    # ── Filas de encabezado ───────────────────────────────────────────
    ROW_LOGO  = 1
    ROW_TITLE = 2
    ROW_TIPO  = 3
    ROW_COMP  = 4
    ROW_RA    = 5
    ROW_COUNT = 6
    ROW_DATOS = 7

    # Fila 1: logo UV
    ws.row_dimensions[ROW_LOGO].height = 50
    if LOGO_PATH.exists():
        img = XLImage(str(LOGO_PATH))
        img.height = 65
        img.width  = int(img.height * (img.width / img.height)) if img.width else 200
        # Calcular ancho proporcional desde las dimensiones reales de la imagen
        try:
            from PIL import Image as PILImage
            with PILImage.open(str(LOGO_PATH)) as pil:
                w_px, h_px = pil.size
            img.width = int(65 * w_px / h_px)
        except Exception:
            img.width = 200
        img.anchor = "A1"
        ws.add_image(img)

    # Fila 2: título del documento
    ws.merge_cells(start_row=ROW_TITLE, start_column=1,
                   end_row=ROW_TITLE, end_column=total_cols)
    _write(ws, ROW_TITLE, 1,
           "MATRIZ DE TRIBUTACIÓN DE COMPETENCIAS — INGENIERÍA CIVIL MATEMÁTICA, PLAN 2025",
           font=_font(bold=True, size=13, color=HEX["licenciatura"]),
           align=_align(h="center"))
    ws.row_dimensions[ROW_TITLE].height = 22

    # Merge verticales cols A-I (filas 2-3 y 2-4 según columna)
    for c in [COL_NIVEL, COL_SEM, COL_COD, COL_NOM, COL_CRED,
              COL_TOT, COL_LIC, COL_TIT, COL_SELLO]:
        ws.merge_cells(start_row=ROW_TIPO, start_column=c,
                       end_row=ROW_COMP,   end_column=c)

    # Etiquetas de las columnas fijas (fila 2, se extiende hasta fila 3 por merge)
    for c, lbl, fillhex in [
        (COL_NIVEL, "Nivel",   HEX_GRIS),
        (COL_SEM,   "Sem",     HEX_GRIS),
        (COL_COD,   "Código",  HEX_GRIS),
        (COL_NOM,   "Nombre de Asignatura", HEX_GRIS),
        (COL_CRED,  "Créd.",   HEX_GRIS),
        (COL_TOT,   "Total",   HEX_GRIS),
        (COL_LIC,   "Lic.",    HEX_CLARO["licenciatura"]),
        (COL_TIT,   "Tít.",    HEX_CLARO["titulo"]),
        (COL_SELLO, "Sello",   HEX_CLARO["sello_uv"]),
    ]:
        _write(ws, ROW_TIPO, c, lbl,
               font=_font(bold=True, size=9),
               fill=_fill(fillhex),
               align=_align(h="center"),
               border=_border_thin())

    # Fila 2: bloques de tipo sobre columnas RA
    for comp_tipo, comp_list in tipo_comps.items():
        cols = []
        for _, ra_ids in comp_list:
            cols += [ra_to_col[r] for r in ra_ids]
        c1, c2 = min(cols), max(cols)
        if c1 != c2:
            ws.merge_cells(start_row=ROW_TIPO, start_column=c1,
                           end_row=ROW_TIPO,   end_column=c2)
        _write(ws, ROW_TIPO, c1, LABEL[comp_tipo],
               font=_font(bold=True, size=9, color=HEX_BLANCO),
               fill=_fill(HEX[comp_tipo]),
               align=_align(h="center"),
               border=_border_thin())

    # Fila 3: código de competencia (merge por sus RAs)
    for (comp_cod, comp_tipo), ra_ids in comp_ras.items():
        cols = [ra_to_col[r] for r in ra_ids]
        c1, c2 = min(cols), max(cols)
        if c1 != c2:
            ws.merge_cells(start_row=ROW_COMP, start_column=c1,
                           end_row=ROW_COMP,   end_column=c2)
        _write(ws, ROW_COMP, c1, comp_cod,
               font=_font(bold=True, size=9, color=HEX_BLANCO),
               fill=_fill(HEX[comp_tipo]),
               align=_align(h="center"),
               border=_border_thin())

    # Fila 4: código RA rotado 90°
    ws.row_dimensions[ROW_RA].height = 80
    for c in range(1, COL_RA_INI):
        ws.cell(ROW_RA, c).border = _border_thin()
    for ra_id, ccomp, nivel, cod_ra, comp_cod, comp_tipo in ra_list:
        c = ra_to_col[ra_id]
        bg = _bg_ra(comp_cod, comp_tipo, nivel)
        _write(ws, ROW_RA, c, ccomp,
               font=_font(size=8, color=HEX[comp_tipo]),
               fill=_fill(bg),
               align=_align(h="center", v="bottom", rotation=90),
               border=_border_thin())

    # Fila 5: COUNTIF por columna RA
    ws.row_dimensions[ROW_COUNT].height = 16
    UF = ROW_DATOS + len(asigs) - 1

    for c in range(1, COL_RA_INI):
        _write(ws, ROW_COUNT, c, "",
               fill=_fill(HEX_GRIS), border=_border_thin())
    _write(ws, ROW_COUNT, COL_NOM, "Subtotales por RA →",
           font=_font(bold=True, size=9),
           fill=_fill(HEX_GRIS), align=_align(h="right"),
           border=_border_thin())

    for ra_id, ccomp, nivel, cod_ra, comp_cod, comp_tipo in ra_list:
        c = ra_to_col[ra_id]
        cl = get_column_letter(c)
        bg = _bg_ra(comp_cod, comp_tipo, nivel)
        _write(ws, ROW_COUNT, c,
               f'=COUNTIF({cl}{ROW_DATOS}:{cl}{UF},"X")',
               font=_font(bold=True, size=9),
               fill=_fill(bg), align=_align(h="center"),
               border=_border_thin())

    # ── Filas de asignaturas ──────────────────────────────────────────
    sem_filas = OrderedDict()
    for idx, (asig_id, cod, nom, sem, cred) in enumerate(asigs):
        sem_filas.setdefault(sem, []).append(ROW_DATOS + idx)

    sem_actual = None
    for idx, (asig_id, cod, nom, sem, cred) in enumerate(asigs):
        fila = ROW_DATOS + idx
        ws.row_dimensions[fila].height = 18

        if sem != sem_actual:
            sem_actual = sem
            filas_sem = sem_filas[sem]
            f1, f2 = filas_sem[0], filas_sem[-1]
            niv = _nivel_academico(sem)
            if f1 != f2:
                ws.merge_cells(start_row=f1, start_column=COL_NIVEL,
                               end_row=f2,   end_column=COL_NIVEL)
                ws.merge_cells(start_row=f1, start_column=COL_SEM,
                               end_row=f2,   end_column=COL_SEM)
            _write(ws, f1, COL_NIVEL, niv,
                   font=_font(bold=True, size=14, color=HEX["licenciatura"]),
                   fill=_fill(HEX_CLARO["licenciatura"]),
                   align=_align(h="center", v="center", rotation=90),
                   border=_border_thin())
            _write(ws, f1, COL_SEM, sem,
                   font=_font(bold=True, size=12),
                   fill=_fill(HEX_CLARO["licenciatura"]),
                   align=_align(h="center"),
                   border=_border_thin())

        _write(ws, fila, COL_COD, cod,
               font=_font(bold=True, size=10),
               align=_align(h="center"), border=_border_thin())

        _write(ws, fila, COL_NOM, nom,
               font=_font(size=10),
               align=_align(h="left", wrap=True), border=_border_thin())

        _write(ws, fila, COL_CRED, cred,
               font=_font(size=9), fill=_fill(HEX_GRIS),
               align=_align(h="center"), border=_border_thin())

        gl = get_column_letter(COL_LIC)
        sl = get_column_letter(COL_SELLO)
        _write(ws, fila, COL_TOT,
               f"=SUM({gl}{fila}:{sl}{fila})",
               font=_font(bold=True, size=10),
               align=_align(h="center"), border=_border_thin())

        for col_r, comp_tipo in [
            (COL_LIC,   "licenciatura"),
            (COL_TIT,   "titulo"),
            (COL_SELLO, "sello_uv"),
        ]:
            rng = tipo_col_range.get(comp_tipo)
            if rng:
                c1l = get_column_letter(rng[0])
                c2l = get_column_letter(rng[1])
                _write(ws, fila, col_r,
                       f'=COUNTIF({c1l}{fila}:{c2l}{fila},"X")',
                       font=_font(size=10),
                       align=_align(h="center"), border=_border_thin())
            else:
                _write(ws, fila, col_r, 0,
                       font=_font(size=10),
                       align=_align(h="center"), border=_border_thin())

        for ra_id, ccomp, nivel, cod_ra, comp_cod, comp_tipo in ra_list:
            c = ra_to_col[ra_id]
            bg = _bg_ra(comp_cod, comp_tipo, nivel)
            if (asig_id, ra_id) in tribs:
                _write(ws, fila, c, "X",
                       font=_font(bold=True, size=10, color=HEX[comp_tipo]),
                       fill=_fill(bg),
                       align=_align(h="center"), border=_border_thin())
            else:
                _write(ws, fila, c, "",
                       fill=_fill(bg),
                       border=_border_thin())

    # Congelar encabezados
    ws.freeze_panes = ws.cell(ROW_DATOS, COL_RA_INI)

    # ── Guardar ───────────────────────────────────────────────────────
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if salida is None:
        salida = str(OUT_DIR / f"matriz_competencias_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
    wb.save(salida)
    print(f"✓ Matriz generada: {salida}")
    print(f"  {len(asigs)} asignaturas × {len(ra_list)} RAs")
    return salida


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--salida", default=None)
    a = p.parse_args()
    generar_matriz(salida=a.salida)
